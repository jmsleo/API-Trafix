"""Render finance reports as CSV / Excel / PDF downloads.

The route layer builds an :class:`ExportDocument` (title, period, summary
lines and one or more pre-formatted tables) from the existing report CRUD
functions; this module turns that document into bytes for each format:

* CSV — stdlib :mod:`csv`, streamed, utf-8 with BOM so Excel opens it
  without an import wizard. All sections are written sequentially.
* XLSX — openpyxl in write-only mode (bounded memory). One sheet per table
  plus a leading "Ringkasan" sheet holding title/period/summary lines.
* PDF — reportlab platypus: landscape A4, site name + period header, zebra
  tables with repeating header rows and page numbers in the footer.

Values inside tables are display-ready primitives (strings for money and
timestamps so all three formats look identical); counts stay integers so
Excel can sort them.
"""

from __future__ import annotations

import csv
import io
import re
from collections.abc import Iterator
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from enum import Enum
from urllib.parse import quote

from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from api_trafix.config.settings import get_settings

WIB = timezone(timedelta(hours=7))

#: Hard ceiling on exported rows — keeps PDF/XLSX generation bounded and
#: nudges users toward narrower date filters instead of full-history dumps.
EXPORT_MAX_ROWS = 10_000


class ExportFormat(str, Enum):
    CSV = "csv"
    XLSX = "xlsx"
    PDF = "pdf"


MEDIA_TYPES = {
    ExportFormat.CSV: "text/csv; charset=utf-8",
    ExportFormat.XLSX: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ExportFormat.PDF: "application/pdf",
}


# ---------------------------------------------------------------------------
# Document model
# ---------------------------------------------------------------------------


@dataclass
class ExportTable:
    """One rendered table: sheet in Excel / section heading elsewhere."""

    title: str
    columns: list[str]
    rows: list[list[object]] = field(default_factory=list)


@dataclass
class ExportDocument:
    filename_base: str  # ASCII slug, e.g. "laporan-transaksi"
    title: str
    period: str | None = None
    summary_lines: list[str] = field(default_factory=list)
    tables: list[ExportTable] = field(default_factory=list)

    @property
    def generated_at(self) -> str:
        return datetime.now(WIB).strftime("%d/%m/%Y %H:%M")


# ---------------------------------------------------------------------------
# Formatting helpers used by the document builders
# ---------------------------------------------------------------------------


def fmt_dt(value: datetime | None) -> str:
    """UTC timestamp -> WIB display string."""
    if value is None:
        return "-"
    if isinstance(value, datetime):
        return value.astimezone(WIB).strftime("%d/%m/%Y %H:%M")
    return str(value)


def fmt_date(value: date | None) -> str:
    if value is None:
        return "-"
    return value.strftime("%d/%m/%Y")


def rp(value: int | float | None) -> str:
    """Rupiah amount with Indonesian thousand separators ("Rp4.000")."""
    if value is None:
        return "-"
    return "Rp" + f"{int(value):,}".replace(",", ".")


def period_label(start_date: date | None, end_date: date | None) -> str | None:
    if start_date is None and end_date is None:
        return None
    start = fmt_date(start_date) if start_date else "…"
    end = fmt_date(end_date) if end_date else "…"
    return f"{start} - {end}"


def duration_minutes(entry: datetime | None, exit_at: datetime | None) -> int | None:
    if entry is None or exit_at is None:
        return None
    return max(0, int((exit_at - entry).total_seconds() // 60))


def enum_value(value) -> str | None:
    return value.value if hasattr(value, "value") else value


# ---------------------------------------------------------------------------
# Generic renderers
# ---------------------------------------------------------------------------


def build_export_response(doc: ExportDocument, export_format: ExportFormat) -> Response:
    """Dispatch to the right renderer and wrap in a download response."""
    filename = f"{doc.filename_base}.{export_format.value}"
    disposition = f"attachment; filename=\"{filename}\"; filename*=UTF-8''{quote(filename)}"

    if export_format is ExportFormat.CSV:
        return StreamingResponse(
            render_csv(doc),
            media_type=MEDIA_TYPES[export_format],
            headers={"Content-Disposition": disposition},
        )
    renderer = render_xlsx if export_format is ExportFormat.XLSX else render_pdf
    return Response(
        content=renderer(doc),
        media_type=MEDIA_TYPES[export_format],
        headers={"Content-Disposition": disposition},
    )


def _meta_lines(doc: ExportDocument) -> list[tuple[str, bool]]:
    """(line, bold) pairs describing the header block of every format."""
    lines = [(doc.title, True)]
    if doc.period:
        lines.append((f"Periode: {doc.period}", False))
    site_name = get_settings().site_name
    if site_name:
        lines.append((f"Lokasi: {site_name}", False))
    lines.extend((line, False) for line in doc.summary_lines)
    lines.append((f"Dibuat: {doc.generated_at} WIB", False))
    return lines


def render_csv(doc: ExportDocument) -> Iterator[bytes]:
    # utf-8-sig BOM up front so Windows Excel detects the encoding.
    yield b"\xef\xbb\xbf"

    buf = io.StringIO(newline="")
    writer = csv.writer(buf)

    def flush() -> Iterator[bytes]:
        yield buf.getvalue().encode("utf-8")
        buf.seek(0)
        buf.truncate()

    for line, _bold in _meta_lines(doc):
        writer.writerow([line])
        yield from flush()

    for table in doc.tables:
        writer.writerow([])
        writer.writerow([table.title])
        writer.writerow(table.columns)
        for row in table.rows:
            writer.writerow(["" if v is None else v for v in row])
        yield from flush()


def _sheet_name(base: str, used: set[str]) -> str:
    name = re.sub(r"[\[\]\:\*\?\/\\]", "-", base)[:31].strip() or "Sheet"
    candidate = name
    counter = 2
    while candidate in used:
        suffix = f" ({counter})"
        candidate = name[: 31 - len(suffix)] + suffix
        counter += 1
    used.add(candidate)
    return candidate


def render_xlsx(doc: ExportDocument) -> bytes:
    wb = Workbook(write_only=True)
    bold = Font(bold=True)
    used_sheets: set[str] = set()

    meta = wb.create_sheet(_sheet_name("Ringkasan", used_sheets))
    meta_width = 10
    for line, is_bold in _meta_lines(doc):
        cell = WriteOnlyCell(meta, value=line)
        cell.font = bold if is_bold else Font()
        meta.append([cell])
        meta_width = max(meta_width, len(line))
    # Column widths are safe to assign any time before save(): write-only
    # sheets buffer rows in memory and serialize everything at once.
    meta.column_dimensions["A"].width = min(meta_width + 2, 80)

    for table in doc.tables:
        ws = wb.create_sheet(_sheet_name(table.title, used_sheets))
        col_widths = [len(str(c)) for c in table.columns]
        ws.append([_bold_cell(ws, column, bold) for column in table.columns])
        for row in table.rows:
            ws.append([v if v is not None else "" for v in row])
            for i, v in enumerate(row):
                col_widths[i] = max(col_widths[i], len(str(v if v is not None else "")))
        for i, width in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = min(max(width + 2, 10), 60)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _bold_cell(ws, value: str, font: Font) -> WriteOnlyCell:
    cell = WriteOnlyCell(ws, value=value)
    cell.font = font
    return cell


_PDF_HEADER_BG = colors.HexColor("#BF8F51")
_PDF_ZEBRA_BG = colors.HexColor("#F5EFE6")
_PDF_GRID = colors.HexColor("#D8CDBB")


def render_pdf(doc: ExportDocument) -> bytes:
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ExportTitle", parent=styles["Title"], fontSize=16, spaceAfter=2
    )
    subtitle_style = ParagraphStyle(
        "ExportSubtitle", parent=styles["Normal"], fontSize=9, textColor=colors.grey
    )
    section_style = ParagraphStyle(
        "ExportSection", parent=styles["Heading3"], fontSize=11, spaceBefore=10, spaceAfter=4
    )
    cell_style = ParagraphStyle("ExportCell", parent=styles["Normal"], fontSize=7.5, leading=9)

    page = landscape(A4)
    margin = 36

    def footer(canvas, _doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawCentredString(page[0] / 2, 14, f"Halaman {canvas.getPageNumber()}")
        canvas.restoreState()

    bio = io.BytesIO()
    pdf_doc = SimpleDocTemplate(
        bio,
        pagesize=page,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin + 6,
        title=doc.title,
    )

    meta_bits = []
    site_name = get_settings().site_name
    if site_name:
        meta_bits.append(site_name)
    meta_bits.append(f"Periode: {doc.period}" if doc.period else "Semua Periode")
    meta_bits.append(f"Dibuat {doc.generated_at} WIB")

    story: list[object] = [
        Paragraph(doc.title, title_style),
        Paragraph(" &nbsp;|&nbsp; ".join(meta_bits), subtitle_style),
    ]
    for line in doc.summary_lines:
        story.append(Paragraph(line, subtitle_style))

    available_width = page[0] - 2 * margin

    for table in doc.tables:
        story.append(Spacer(1, 12))
        story.append(Paragraph(table.title, section_style))
        ncols = max(len(table.columns), 1)
        body = [[Paragraph(str(c), cell_style) for c in table.columns]]
        for row in table.rows:
            body.append(
                [Paragraph("-" if v is None else str(v), cell_style) for v in row]
            )
        platypus_table = Table(
            body, colWidths=[available_width / ncols] * ncols, repeatRows=1
        )
        platypus_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), _PDF_HEADER_BG),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _PDF_ZEBRA_BG]),
                    ("GRID", (0, 0), (-1, -1), 0.4, _PDF_GRID),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(platypus_table)

    pdf_doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return bio.getvalue()
